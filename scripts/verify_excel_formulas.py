"""
Verify Excel formulas in the Customer Churn workbook.
"""
import openpyxl

wb = openpyxl.load_workbook('excel/churn_analysis_summary.xlsx')

print('=' * 70)
print('EXCEL FORMULA VERIFICATION - Customer Churn Analysis')
print('=' * 70)

print(f'\nSheets: {wb.sheetnames}')

# Count formulas per sheet
print('\n[FORMULA COUNT PER SHEET]')
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    formula_count = 0
    sample_formulas = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and str(cell.value).startswith('='):
                formula_count += 1
                if len(sample_formulas) < 2:
                    sample_formulas.append(f'{cell.coordinate}: {cell.value}')
    
    status = '[OK]' if formula_count > 0 else '[--]'
    print(f'   {status} {sheet_name}: {formula_count} formulas')
    for f in sample_formulas:
        print(f'       Example: {f}')

# Total
total_formulas = 0
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and str(cell.value).startswith('='):
                total_formulas += 1

print('\n' + '=' * 70)
print(f'Total formulas in workbook: {total_formulas}')
print('=' * 70)
