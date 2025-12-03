"""
Customer Churn Analysis - Enhanced Excel Workbook Generator
============================================================
Creates a professional Excel workbook with DYNAMIC formulas.

Author: Md Imran Hossain
Date: November 2025
"""

import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from datetime import datetime

# =============================================================================
# PATHS
# =============================================================================
PROJECT_ROOT = Path(__file__).parent.parent
DATA_PATH = PROJECT_ROOT / "data" / "processed" / "customers_cleaned.csv"
OUTPUT_PATH = PROJECT_ROOT / "excel" / "churn_analysis_dynamic.xlsx"

# =============================================================================
# STYLES
# =============================================================================
HEADER_FILL = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
KPI_FILL = PatternFill(start_color="FBE5D6", end_color="FBE5D6", fill_type="solid")
CHURN_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
RETAIN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

HEADER_FONT = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="C65911")
KPI_FONT = Font(name="Calibri", size=14, bold=True)
MONEY_FONT = Font(name="Calibri", size=14, bold=True, color="006600")

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

CENTER_ALIGN = Alignment(horizontal='center', vertical='center')


def auto_adjust_columns(ws):
    """Auto-adjust column widths."""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 40)


def create_dashboard_sheet(wb, df):
    """Create executive dashboard with KPIs."""
    ws = wb.create_sheet("Dashboard", 0)
    
    # Title
    ws['A1'] = "CUSTOMER CHURN ANALYSIS DASHBOARD"
    ws['A1'].font = Font(name="Calibri", size=24, bold=True, color="C65911")
    ws.merge_cells('A1:G1')
    
    ws['A2'] = f"Analysis Date: {datetime.now().strftime('%B %d, %Y')}"
    ws['A2'].font = Font(italic=True)
    
    # KPI Section - Using formulas referencing Raw_Data
    ws['A4'] = "KEY PERFORMANCE INDICATORS"
    ws['A4'].font = TITLE_FONT
    
    # KPI Grid
    kpis = [
        ("A6", "Total Customers", "B6", "=COUNTA(Raw_Data!A2:A7044)"),
        ("C6", "Churned", "D6", "=COUNTIF(Raw_Data!U2:U7044,\"Yes\")"),
        ("E6", "Retained", "F6", "=COUNTIF(Raw_Data!U2:U7044,\"No\")"),
        ("A8", "Churn Rate", "B8", "=D6/B6"),
        ("C8", "Retention Rate", "D8", "=F6/B6"),
        ("E8", "Avg Monthly Charges", "F8", "=AVERAGE(Raw_Data!S2:S7044)"),
        ("A10", "Total Monthly Revenue", "B10", "=SUM(Raw_Data!S2:S7044)"),
        ("C10", "Revenue at Risk", "D10", "=SUMIF(Raw_Data!U2:U7044,\"Yes\",Raw_Data!S2:S7044)"),
        ("E10", "Annual Risk ($)", "F10", "=D10*12"),
    ]
    
    for label_cell, label, value_cell, formula in kpis:
        ws[label_cell] = label
        ws[label_cell].font = Font(bold=True, size=10)
        ws[label_cell].fill = SUBHEADER_FILL
        ws[label_cell].font = HEADER_FONT
        ws[label_cell].border = THIN_BORDER
        
        ws[value_cell] = formula
        ws[value_cell].font = KPI_FONT
        ws[value_cell].fill = KPI_FILL
        ws[value_cell].border = THIN_BORDER
        ws[value_cell].alignment = CENTER_ALIGN
        
        # Format percentages and currency
        if "Rate" in label:
            ws[value_cell].number_format = '0.00%'
        elif "Revenue" in label or "Risk" in label or "Charges" in label:
            ws[value_cell].number_format = '$#,##0.00'
    
    # Insight Box
    ws['A13'] = "KEY INSIGHT"
    ws['A13'].font = TITLE_FONT
    
    ws['A14'] = "Revenue at Risk Formula: =SUMIF(Churn='Yes', MonthlyCharges) * 12"
    ws['A15'] = "This represents annual revenue that could be lost if all at-risk customers churn."
    
    auto_adjust_columns(ws)
    return ws


def create_churn_by_contract_sheet(wb, df):
    """Churn analysis by contract type with formulas."""
    ws = wb.create_sheet("Churn_by_Contract")
    
    ws['A1'] = "CHURN ANALYSIS BY CONTRACT TYPE"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:F1')
    
    # Headers
    headers = ['Contract Type', 'Total', 'Churned', 'Retained', 'Churn Rate', 'Revenue at Risk']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    
    # Contract types
    contracts = ['Month-to-month', 'One year', 'Two year']
    
    for i, contract in enumerate(contracts):
        row = i + 4
        
        # Contract Type
        ws.cell(row=row, column=1, value=contract).border = THIN_BORDER
        
        # Total (COUNTIF)
        ws.cell(row=row, column=2, 
                value=f'=COUNTIF(Raw_Data!O2:O7044,"{contract}")').border = THIN_BORDER
        
        # Churned (COUNTIFS)
        ws.cell(row=row, column=3,
                value=f'=COUNTIFS(Raw_Data!O2:O7044,"{contract}",Raw_Data!U2:U7044,"Yes")').border = THIN_BORDER
        
        # Retained (formula)
        ws.cell(row=row, column=4, value=f'=B{row}-C{row}').border = THIN_BORDER
        
        # Churn Rate (formula)
        churn_cell = ws.cell(row=row, column=5, value=f'=C{row}/B{row}')
        churn_cell.number_format = '0.00%'
        churn_cell.border = THIN_BORDER
        
        # Revenue at Risk (SUMIFS)
        rev_cell = ws.cell(row=row, column=6,
                          value=f'=SUMIFS(Raw_Data!S2:S7044,Raw_Data!O2:O7044,"{contract}",Raw_Data!U2:U7044,"Yes")*12')
        rev_cell.number_format = '$#,##0'
        rev_cell.border = THIN_BORDER
    
    # Total row
    row = 7
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=1).border = THIN_BORDER
    ws.cell(row=row, column=2, value="=SUM(B4:B6)").border = THIN_BORDER
    ws.cell(row=row, column=2).font = Font(bold=True)
    ws.cell(row=row, column=3, value="=SUM(C4:C6)").border = THIN_BORDER
    ws.cell(row=row, column=3).font = Font(bold=True)
    ws.cell(row=row, column=4, value="=SUM(D4:D6)").border = THIN_BORDER
    ws.cell(row=row, column=4).font = Font(bold=True)
    ws.cell(row=row, column=5, value="=C7/B7").number_format = '0.00%'
    ws.cell(row=row, column=5).border = THIN_BORDER
    ws.cell(row=row, column=5).font = Font(bold=True)
    ws.cell(row=row, column=6, value="=SUM(F4:F6)").number_format = '$#,##0'
    ws.cell(row=row, column=6).border = THIN_BORDER
    ws.cell(row=row, column=6).font = Font(bold=True)
    
    # Add color scale to churn rate
    ws.conditional_formatting.add(
        'E4:E6',
        ColorScaleRule(start_type='min', start_color='63BE7B',
                      end_type='max', end_color='F8696B')
    )
    
    # Add bar chart
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Churn Rate by Contract Type"
    chart.y_axis.title = "Churn Rate"
    
    data = Reference(ws, min_col=5, min_row=3, max_row=6)
    cats = Reference(ws, min_col=1, min_row=4, max_row=6)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 12
    chart.height = 8
    
    ws.add_chart(chart, "H3")
    
    auto_adjust_columns(ws)
    return ws


def create_churn_by_payment_sheet(wb, df):
    """Churn analysis by payment method."""
    ws = wb.create_sheet("Churn_by_Payment")
    
    ws['A1'] = "CHURN ANALYSIS BY PAYMENT METHOD"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:F1')
    
    headers = ['Payment Method', 'Total', 'Churned', 'Retained', 'Churn Rate', 'Revenue at Risk']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    
    payments = ['Electronic check', 'Mailed check', 'Bank transfer (automatic)', 'Credit card (automatic)']
    
    for i, payment in enumerate(payments):
        row = i + 4
        
        ws.cell(row=row, column=1, value=payment).border = THIN_BORDER
        ws.cell(row=row, column=2,
                value=f'=COUNTIF(Raw_Data!P2:P7044,"{payment}")').border = THIN_BORDER
        ws.cell(row=row, column=3,
                value=f'=COUNTIFS(Raw_Data!P2:P7044,"{payment}",Raw_Data!U2:U7044,"Yes")').border = THIN_BORDER
        ws.cell(row=row, column=4, value=f'=B{row}-C{row}').border = THIN_BORDER
        
        churn_cell = ws.cell(row=row, column=5, value=f'=C{row}/B{row}')
        churn_cell.number_format = '0.00%'
        churn_cell.border = THIN_BORDER
        
        rev_cell = ws.cell(row=row, column=6,
                          value=f'=SUMIFS(Raw_Data!S2:S7044,Raw_Data!P2:P7044,"{payment}",Raw_Data!U2:U7044,"Yes")*12')
        rev_cell.number_format = '$#,##0'
        rev_cell.border = THIN_BORDER
    
    # Add color scale
    ws.conditional_formatting.add(
        'E4:E7',
        ColorScaleRule(start_type='min', start_color='63BE7B',
                      end_type='max', end_color='F8696B')
    )
    
    auto_adjust_columns(ws)
    return ws


def create_tenure_analysis_sheet(wb, df):
    """Churn analysis by tenure buckets."""
    ws = wb.create_sheet("Churn_by_Tenure")
    
    ws['A1'] = "CHURN ANALYSIS BY CUSTOMER TENURE"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:F1')
    
    headers = ['Tenure Bucket', 'Min Months', 'Max Months', 'Customers', 'Churned', 'Churn Rate']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    
    # Tenure buckets
    buckets = [
        ('0-12 months', 0, 12),
        ('13-24 months', 13, 24),
        ('25-36 months', 25, 36),
        ('37-48 months', 37, 48),
        ('49-60 months', 49, 60),
        ('61-72 months', 61, 72),
    ]
    
    for i, (label, min_m, max_m) in enumerate(buckets):
        row = i + 4
        
        ws.cell(row=row, column=1, value=label).border = THIN_BORDER
        ws.cell(row=row, column=2, value=min_m).border = THIN_BORDER
        ws.cell(row=row, column=3, value=max_m).border = THIN_BORDER
        
        # Customers in range (COUNTIFS)
        ws.cell(row=row, column=4,
                value=f'=COUNTIFS(Raw_Data!F2:F7044,">="&B{row},Raw_Data!F2:F7044,"<="&C{row})').border = THIN_BORDER
        
        # Churned in range
        ws.cell(row=row, column=5,
                value=f'=COUNTIFS(Raw_Data!F2:F7044,">="&B{row},Raw_Data!F2:F7044,"<="&C{row},Raw_Data!U2:U7044,"Yes")').border = THIN_BORDER
        
        # Churn Rate
        churn_cell = ws.cell(row=row, column=6, value=f'=IF(D{row}=0,0,E{row}/D{row})')
        churn_cell.number_format = '0.00%'
        churn_cell.border = THIN_BORDER
    
    # Color scale
    ws.conditional_formatting.add(
        'F4:F9',
        ColorScaleRule(start_type='min', start_color='63BE7B',
                      end_type='max', end_color='F8696B')
    )
    
    # Key insight
    ws['A11'] = "KEY INSIGHT:"
    ws['A11'].font = Font(bold=True)
    ws['A12'] = "New customers (0-12 months) have the highest churn risk."
    ws['A13'] = "Focus retention efforts on the first year of customer relationship."
    
    # Add line chart
    chart = LineChart()
    chart.style = 10
    chart.title = "Churn Rate by Tenure"
    chart.y_axis.title = "Churn Rate"
    chart.x_axis.title = "Tenure Bucket"
    
    data = Reference(ws, min_col=6, min_row=3, max_row=9)
    cats = Reference(ws, min_col=1, min_row=4, max_row=9)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 14
    chart.height = 8
    
    ws.add_chart(chart, "H3")
    
    auto_adjust_columns(ws)
    return ws


def create_service_impact_sheet(wb, df):
    """Analyze churn by service subscriptions."""
    ws = wb.create_sheet("Service_Impact")
    
    ws['A1'] = "CHURN IMPACT BY SERVICE SUBSCRIPTIONS"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:F1')
    
    headers = ['Service', 'Has Service', 'Churned', 'Churn Rate', 'Without Service', 'Churn Rate (No)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    
    # Services with their column letters in Raw_Data
    services = [
        ('Online Security', 'J'),
        ('Online Backup', 'K'),
        ('Device Protection', 'L'),
        ('Tech Support', 'M'),
        ('Streaming TV', 'N'),
    ]
    
    for i, (service, col_letter) in enumerate(services):
        row = i + 4
        
        ws.cell(row=row, column=1, value=service).border = THIN_BORDER
        
        # Has Service
        ws.cell(row=row, column=2,
                value=f'=COUNTIF(Raw_Data!{col_letter}2:{col_letter}7044,"Yes")').border = THIN_BORDER
        
        # Churned with service
        ws.cell(row=row, column=3,
                value=f'=COUNTIFS(Raw_Data!{col_letter}2:{col_letter}7044,"Yes",Raw_Data!U2:U7044,"Yes")').border = THIN_BORDER
        
        # Churn Rate with service
        churn1 = ws.cell(row=row, column=4, value=f'=IF(B{row}=0,0,C{row}/B{row})')
        churn1.number_format = '0.00%'
        churn1.border = THIN_BORDER
        
        # Without Service
        ws.cell(row=row, column=5,
                value=f'=COUNTIF(Raw_Data!{col_letter}2:{col_letter}7044,"No")').border = THIN_BORDER
        
        # Churn Rate without service
        ws.cell(row=row, column=6,
                value=f'=COUNTIFS(Raw_Data!{col_letter}2:{col_letter}7044,"No",Raw_Data!U2:U7044,"Yes")/E{row}')
        ws.cell(row=row, column=6).number_format = '0.00%'
        ws.cell(row=row, column=6).border = THIN_BORDER
    
    # Key insight
    ws['A10'] = "INSIGHT: Customers WITHOUT these services have higher churn rates."
    ws['A10'].font = Font(bold=True, color="C65911")
    ws['A11'] = "Recommendation: Upsell protective services to reduce churn risk."
    
    auto_adjust_columns(ws)
    return ws


def create_revenue_analysis_sheet(wb, df):
    """Revenue analysis and projections."""
    ws = wb.create_sheet("Revenue_Analysis")
    
    ws['A1'] = "REVENUE AT RISK ANALYSIS"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:D1')
    
    # Revenue metrics
    metrics = [
        ('Total Monthly Revenue', '=SUM(Raw_Data!S2:S7044)', '$#,##0'),
        ('Revenue from Churned', '=SUMIF(Raw_Data!U2:U7044,"Yes",Raw_Data!S2:S7044)', '$#,##0'),
        ('Revenue from Retained', '=SUMIF(Raw_Data!U2:U7044,"No",Raw_Data!S2:S7044)', '$#,##0'),
        ('Monthly Revenue at Risk', '=B5', '$#,##0'),
        ('Annual Revenue at Risk', '=B7*12', '$#,##0'),
        ('Avg Charges (Churned)', '=AVERAGEIF(Raw_Data!U2:U7044,"Yes",Raw_Data!S2:S7044)', '$#,##0.00'),
        ('Avg Charges (Retained)', '=AVERAGEIF(Raw_Data!U2:U7044,"No",Raw_Data!S2:S7044)', '$#,##0.00'),
    ]
    
    for i, (label, formula, fmt) in enumerate(metrics):
        row = i + 4
        ws.cell(row=row, column=1, value=label).border = THIN_BORDER
        ws.cell(row=row, column=1).font = Font(bold=True)
        
        val_cell = ws.cell(row=row, column=2, value=formula)
        val_cell.number_format = fmt
        val_cell.border = THIN_BORDER
        val_cell.fill = KPI_FILL
    
    # Retention scenarios
    ws['A13'] = "RETENTION SCENARIO ANALYSIS"
    ws['A13'].font = TITLE_FONT
    
    ws['A15'] = "If we retain"
    ws['B15'] = "We save annually"
    ws['A15'].font = Font(bold=True)
    ws['B15'].font = Font(bold=True)
    
    scenarios = [10, 25, 50, 75, 100]
    for i, pct in enumerate(scenarios):
        row = 16 + i
        ws.cell(row=row, column=1, value=f'{pct}% of at-risk').border = THIN_BORDER
        
        save_cell = ws.cell(row=row, column=2, value=f'=B8*{pct/100}')
        save_cell.number_format = '$#,##0'
        save_cell.border = THIN_BORDER
        if pct == 50:
            save_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    auto_adjust_columns(ws)
    return ws


def create_raw_data_sheet(wb, df):
    """Create raw data sheet."""
    ws = wb.create_sheet("Raw_Data")
    
    # Write data
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
    
    auto_adjust_columns(ws)
    return ws


def main():
    print("=" * 70)
    print("CUSTOMER CHURN - ENHANCED EXCEL WORKBOOK GENERATOR")
    print("=" * 70)
    
    # Load data
    print("\nLoading data...")
    df = pd.read_csv(DATA_PATH)
    print(f"   Loaded {len(df):,} records")
    
    # Create output directory
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    
    # Create workbook
    print("\nCreating Excel workbook with DYNAMIC formulas...")
    wb = Workbook()
    
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    print("   -> Dashboard (KPI formulas)")
    create_dashboard_sheet(wb, df)
    
    print("   -> Churn by Contract (COUNTIFS, SUMIFS)")
    create_churn_by_contract_sheet(wb, df)
    
    print("   -> Churn by Payment (COUNTIFS, SUMIFS)")
    create_churn_by_payment_sheet(wb, df)
    
    print("   -> Churn by Tenure (range formulas)")
    create_tenure_analysis_sheet(wb, df)
    
    print("   -> Service Impact (cross-analysis)")
    create_service_impact_sheet(wb, df)
    
    print("   -> Revenue Analysis (scenarios)")
    create_revenue_analysis_sheet(wb, df)
    
    print("   -> Raw Data")
    create_raw_data_sheet(wb, df)
    
    # Save
    print(f"\nSaving to: {OUTPUT_PATH}")
    wb.save(OUTPUT_PATH)
    
    # Count formulas
    total_formulas = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and str(cell.value).startswith('='):
                    total_formulas += 1
    
    print("\n" + "=" * 70)
    print("SUCCESS!")
    print("=" * 70)
    print(f"File: {OUTPUT_PATH}")
    print(f"Sheets: {', '.join(wb.sheetnames)}")
    print(f"Total Dynamic Formulas: {total_formulas}")
    print("\nFormula Types Used:")
    print("   - COUNTIF / COUNTIFS (segmentation)")
    print("   - SUMIF / SUMIFS (revenue calculations)")
    print("   - AVERAGEIF (average by segment)")
    print("   - Percentage calculations")
    print("   - Scenario projections")


if __name__ == "__main__":
    main()
